import pandas as pd
import numpy as np
import os
from scipy.stats import ttest_ind
from scipy.stats import chi2_contingency
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from datetime import datetime

import rpy2.robjects as ro
from rpy2.robjects import numpy2ri
numpy2ri.activate()

import re

########################################################################################################################################
###### STATISTICS ######
decimals = 1
## Builds a Table 1 without any comparisons
def Table1_Study_Cohort_Describe(df, Study_Vars, continuous_vars):
    """Generates a descriptive summary table for a study cohort.

    Args:
        df (pd.DataFrame): The study data.
        Study_Vars (list): List of variables to summarize.
        continuous_vars (list): List of continuous variables for statistical summaries.

    Returns:
        pd.DataFrame: A summary table with descriptive statistics.
    """
    
    table = []
    total_n = len(df)
    
    for var in Study_Vars:
        if var not in df.columns:
            table.append([var, "", ""])
        else:
            n_Values = df[var].notna().sum()
            missing_n = total_n - n_Values
            
            if var in continuous_vars:
                varName = f"{var}, mean (range)"
                if total_n > n_Values:
                    varName += f", n = {n_Values}"
                var_list = df[var].dropna()
                mean_val = np.nanmean(var_list)
                min_val = np.nanmin(var_list)
                max_val = np.nanmax(var_list)
                entry = f"{mean_val:.{decimals}f} ({min_val:.{decimals}f}, {max_val:.{decimals}f})"
                table.append([varName, entry, missing_n])
            else:
                # Check if the variable is binary
                if df[var].dtype == 'bool' or set(df[var].dropna().unique()).issubset({0, 1}):
                    # Binary variable
                    n_present = int(df[var].sum())  # Count of 1s
                    n_percent = (float(n_present) / n_Values) * 100
                    varName = f"{var} (%)"
                    entry = f"{n_present} ({n_percent:.{decimals}f}%)"
                    if total_n > n_Values:
                        varName += f", n = {n_Values}"
                    table.append([varName, entry, missing_n])
                else:# df[var].nunique() > 2:
                    # Handle multi-category variables
                    varName = f"{var}"
                    if total_n > n_Values:
                        varName += f", n = {n_Values}"

                    table.append([varName, "", missing_n])
                    for category, count in df[var].value_counts().items():
                        percent = (count / n_Values) * 100
                        table.append([f"      {category} (%)", f"{int(count)} ({percent:.{decimals}f}%)", missing_n])

    
    return pd.DataFrame(table, columns=['Variable', f'Study Cohort N={total_n}', 'N Missing'])

 
def Build_Uni_Comparison_Table(df, independent_var, Study_Vars, continunous_vars, Fisher = False, output_Multi = True):
    table = []
    unique_values = df[independent_var].unique()

    # Check if the variable has exactly two unique values
    if len(unique_values) == 2:
        if set(unique_values) == {0, 1}:
            print("The variable is binary (0 or 1). No dummy column created.")
        else:
            # Create a dummy variable for the second unique value
            second_value = unique_values[1]
            df[second_value] = (df[independent_var] == second_value).astype(int)
            independent_var = second_value
    else:
        print("The variable does not have exactly two unique values.")

    df_comp = df[~df[independent_var].isna()]

    multivariate_vars = []
    Yates_and_Fisher = False
    for var in Study_Vars:
        if var not in df.columns:
            table.append([var, "", "", "", ""])
        elif var != independent_var:
            extend = False
            if var in continunous_vars:
                row, multi = Students_Ttest_Build(df_comp, var, independent_var)
            else:
                contingency = pd.crosstab(df[var], df[independent_var])
                if len(contingency) > 5:
                    Yates_and_Fisher = True
                row, extend, multi = Chi_Test_Build(df_comp, var, contingency, Fisher)
            if extend:
                table.extend(row)
            else:
                table.append(row)

            if multi:
                multivariate_vars.append(var)
    if output_Multi:
        print("R code for multivariate: ")
        print(R_Code_Multivariate(independent_var, multivariate_vars))
        
    N_dep_var0 = len(df_comp[df_comp[independent_var]==0])
    N_dep_var1 = len(df_comp[df_comp[independent_var]==1])
    Denom_dep_var = df_comp[independent_var].notna().sum()

    correction_str = "†Ch² with Yates correction"
    if Yates_and_Fisher and Fisher:
        correction_str += ", ‡Fisher Exact Test"
    elif Fisher:
        correction_str = "‡Fisher Exact Test"
    
    table.append(["","","",correction_str])

    return pd.DataFrame(table, columns=['Variable', 
                                     f"No {independent_var}\nN={N_dep_var0} ({N_dep_var0/Denom_dep_var*100:.1f}%)" , 
                                     f"{independent_var}\nN={N_dep_var1} ({N_dep_var1/Denom_dep_var*100:.1f}%)",
                                     'p value',
                                     'N Missing'])

## DO NOT CALL - USE Build_Uni_Comparison_Table()
## Performs Chi2 analysis
def Chi_Test_Build(df, dependent_var, contingency, Fisher):
    # Create the contingency table
    dependent_var_name = dependent_var + " (%)"
    if contingency.shape[0] < 2 or contingency.shape[1] < 2:
        # Not enough categories to perform the chi-squared test
        entry_01 = " N/A "
        entry_11 = " N/A "
        p = " N/A "
        return [dependent_var_name, entry_01, entry_11, p, len(df) - df[dependent_var].notna().sum()], False, False
    else:
        # Perform chi-squared test without Yates' correction
        c, p, dof, exp = chi2_contingency(contingency, correction=False)
        
        corrected = ""
        # Apply Yates' correction if any expected frequency is less than 5
        if np.sum(exp < 5) > 0.2 * exp.size:
            if (not Fisher) or (contingency.shape[0] > 5 or contingency.shape[1] > 5):
                c, p, dof, exp = chi2_contingency(contingency, correction=True)
                corrected = "†"
            else:
                r_contingency = ro.r.matrix(np.array(contingency), nrow=contingency.shape[0])
                result = ro.r('fisher.test')(r_contingency)
                p = result.rx2('p.value')[0]
                #odds_ratio = result.rx2('estimate')[0]
                corrected = "‡"
        # Keep track of var for multivariate analysis (p<0.2)    
        multi = p < 0.2

        # Format p-value
        if p < 0.001:
            p = "<0.001"
        else:
            p = str(round(p, 3)) 
        p = p + corrected

        ### Contingency sizes
        if len(set(df[dependent_var].dropna().unique()) - {0, 1}) > 0: #len(contingency) > 2:
            # Calculate the total number of values for each row
            contingency['Total'] = contingency.sum(axis=1)
            
            # Sort rows by the total number of values in descending order
            sorted_contingency = contingency.sort_values(by='Total', ascending=False)
            
            # Drop the 'Total' column after sorting
            sorted_contingency = sorted_contingency.drop(columns='Total')

            # Calculate true sample size excluding NaNs
            true_n = df[dependent_var].notna().sum()

            # Update the dependent_var name to include true sample size if it's different
            if true_n != len(df):
                dependent_var_name += f", n = {true_n}"

            # Build the result table
            to_return = [[dependent_var_name, " ", " ", p, len(df) - true_n]]
            
            for j in range(sorted_contingency.shape[0]):
                temp = ['     ' + str(sorted_contingency.index[j]) + " (%)"]
                for i in range(sorted_contingency.shape[1]):
                    denom = sorted_contingency.iloc[:, i].sum()
                    n = sorted_contingency.iloc[j, i]
                    if len(df) > true_n:
                        denom_str = f"/{denom}"
                        temp.append(f"{n}{denom_str} ({np.round(100 * (n / denom), decimals)}%)")
                    else:
                        temp.append(f"{n} ({np.round(100 * (n / denom), decimals)}%)")
                
                temp.extend([" ", " "])
                to_return.append(temp)
            
            return to_return, True, multi

        else:
            # 2x2 table case
            count_01 = contingency[0][1]
            sum_0 = contingency[0].sum()
            percent_01 = str(round(count_01 / sum_0 * 100, decimals))

            count_11 = contingency[1][1]
            sum_1 = contingency[1].sum()
            percent_11 = str(round(count_11 / sum_1 * 100, decimals))

            entry_01 = str(count_01)
            entry_11 = str(count_11)

            if len(df) > df[dependent_var].notna().sum():
                dependent_var_name += f", n = {df[dependent_var].notna().sum()}"
                entry_01 += f"/{sum_0}"
                entry_11 += f"/{sum_1}"

            entry_01 += f" ({percent_01}%)"
            entry_11 += f" ({percent_11}%)"
            
            return [dependent_var_name, entry_01, entry_11, p, len(df) - df[dependent_var].notna().sum()], False, multi

        #return [dependent_var, entry_01, entry_11, p, len(df) - sum(df[dependent_var].value_counts())], False


def Students_Ttest_Build(df, independent_var, dependent_var):
    list0 = [x for x in df[df[dependent_var]==0][independent_var].tolist() if not np.isnan(x)]
    list1 = [x for x in df[df[dependent_var]==1][independent_var].tolist() if not np.isnan(x)]
    t_stat, p = ttest_ind(list0, list1)

    # Keep track of var for mullivariate analysis (p<0.2)    
    multi = p < 0.2

    if p < 0.001:
        p = "<0.001"
    else:
        p = str(round(p, 3))
    
    independent_var_name = independent_var + ", mean ± SD"
        
    list0_mean = round(np.mean(list0),decimals)
    list0_std = round(np.std(list0),decimals)
    entry_0 = str(list0_mean) + " ± " + str(list0_std)
    
    list1_mean = round(np.mean(list1),decimals)
    list1_std = round(np.std(list1),decimals)
    entry_1 = str(list1_mean) + " ± " + str(list1_std)
    
    
    if len(df) > sum(df[independent_var].value_counts()):
        independent_var_name += ", n = " + str(sum(df[independent_var].value_counts()))
        
    return [independent_var_name, entry_0, entry_1, p, len(df) - sum(df[independent_var].value_counts())], multi



########################################################################################################################################


def R_Code_Multivariate(dependent_var, multivariate_vars):
    pattern = r'[ /\\-_,;\'?]'

    # Replace all delimiters with a period
    code_frag = R_friendly(dependent_var) + " ~ "

    #code_frag = re.sub(pattern, '.', dependent_var) + " ~ "
    for i in range(len(multivariate_vars)):
        if i == 0:
            code_frag += R_friendly(multivariate_vars[i])
        else:
            code_frag += " + " + R_friendly(multivariate_vars[i])
    return code_frag

def R_friendly(var):
    pattern = r'[ /\\-_,;\'?]'
    if isinstance(var, str):
        return re.sub(pattern, '.', var)
    elif isinstance(var, list):
        return [re.sub(pattern, '.', v) for v in var]
    return var


########################################################################################################################################
###### EXPORTING ######
def Export_Univariate_Tables_To_Excel(tables, StudyName):
    file_name = File_Name_Generator(StudyName)[0]

    # Save tables to Excel
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        for table_name, df in tables.items():
            df.to_excel(writer, sheet_name=table_name, index=False)
    
    # Call the formatting function
    apply_formatting_to_excel(file_name)


def Combine_Uni_Multi_Excel_Files(StudyName, delete=False):
    # Generate file names
    all_file_list = File_Name_Generator(StudyName)
    output_file = all_file_list[2]
    file_list = all_file_list[:2]

    # Read and store DataFrames and their original sheet names from each file
    file_sheets = []
    for file in file_list:
        xl = pd.ExcelFile(file)
        sheets = xl.sheet_names
        file_sheets.append(sheets)

    # Create interleaved order of sheets
    interleaved_sheets = get_interleaved_sheets(file_sheets, file_list)

    # Read and collect DataFrames in interleaved order
    all_sheets = []
    for file, sheet_name in interleaved_sheets:
        df = pd.read_excel(file, sheet_name=sheet_name)
        all_sheets.append((sheet_name, df))
    
    # Write all DataFrames to a new Excel file with separate sheets
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, df in all_sheets:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Call the formatting function
    apply_formatting_to_excel(output_file)

    # Optionally delete the original files
    if delete:
        for file in file_list:
            try:
                os.remove(file)
                print(f"Deleted file: {file}")
            except FileNotFoundError:
                print(f"File not found: {file}")
            except PermissionError:
                print(f"Permission denied: {file}")


def get_interleaved_sheets(file_sheets, file_list):
    interleaved_sheets = []
    
    # Unpack the sheets from each file
    sheets_file1, sheets_file2 = file_sheets
    file1, file2 = file_list
    
    # Ensure there are sheets to process
    if not sheets_file1 or not sheets_file2:
        return interleaved_sheets
    
    # Add the first sheet from File 1
    interleaved_sheets.append((file1, sheets_file1.pop(0)))
    
    # Interleave remaining sheets from File 1 and all sheets from File 2
    while sheets_file1 or sheets_file2:
        if sheets_file1:
            interleaved_sheets.append((file1, sheets_file1.pop(0)))
        if sheets_file2:
            interleaved_sheets.append((file2, sheets_file2.pop(0)))
    
    return interleaved_sheets


def File_Name_Generator(file_name):
    date = datetime.now().strftime("%m_%d_%Y")

    names_to_return =[]
    for analysis_type in ['Univariate', "Multivariate", 'Uni_Multi']:
        names_to_return.append(file_name + "_" + analysis_type + "_" + date + ".xlsx")
    return names_to_return




def apply_formatting_to_excel(file_name):
    # Load the workbook and apply formatting
    workbook = load_workbook(file_name)
    
    # Define border styles
    thick_border = Border(
        left=Side(border_style='thick'),
        right=Side(border_style='thick'),
        top=Side(border_style='thick'),
        bottom=Side(border_style='thick')
    )
    
    thin_border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    
    # Define alignment styles
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Define fill color for highlighting
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Apply formatting to each sheet
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Identify the "p value" column index
        p_value_col_idx = None
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value and "p value" in str(cell.value).lower():
                p_value_col_idx = col_idx
                break
        
        # Apply thick border to the entire table
        max_row = sheet.max_row
        max_column = sheet.max_column
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
            for cell in row:
                cell.border = thick_border
        
        # Apply thin border to individual rows
        for row in sheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_column):
            for cell in row:
                cell.border = thin_border
        
        # Center align headers
        for cell in sheet[1]:
            cell.alignment = center_alignment
        
        # Adjust column widths and center align cells except the first column
        for idx, column in enumerate(sheet.columns):
            column_letter = column[0].column_letter
            max_length = 0
            
            for cell in column:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
                
                # Center align the content of all columns except the first one
                if idx > 0:
                    cell.alignment = center_alignment
            
            adjusted_width = max_length + 2  # Adding some extra space for readability
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Highlight "p value" column cells based on condition
        if p_value_col_idx:
            for row in sheet.iter_rows(min_row=2, max_row=max_row, min_col=p_value_col_idx, max_col=p_value_col_idx):
                for cell in row:
                    if cell.value:
                        # Convert cell value to float if it's not "<0.001"
                        if isinstance(cell.value, str):
                            if cell.value == "<0.001":
                                cell.fill = highlight_fill
                            else:
                                try:
                                    value = float(cell.value)
                                    if value < 0.05:
                                        cell.fill = highlight_fill
                                except ValueError:
                                    pass  # Ignore cells that cannot be converted to float
                        elif isinstance(cell.value, (int, float)) and cell.value < 0.05:
                            cell.fill = highlight_fill
    
    # Save the workbook with formatting applied
    workbook.save(file_name)

########################################################################################################################################
###### MISCELLANEOUS ######
def compare_excel_files(file1, file2):
    # Load the Excel files
    xl1 = pd.ExcelFile(file1)
    xl2 = pd.ExcelFile(file2)
    
    # Check if both files have the same sheet names
    if set(xl1.sheet_names) != set(xl2.sheet_names):
        return False, "Sheet names differ between the two files."
    
    # Compare sheets
    for sheet_name in xl1.sheet_names:
        df1 = xl1.parse(sheet_name)
        df2 = xl2.parse(sheet_name)
        
        # Check if DataFrames have the same shape
        if df1.shape != df2.shape:
            return False, f"Sheet '{sheet_name}' has different shapes."
        
        # Check if DataFrames are equal
        if not df1.equals(df2):
            return False, f"Sheet '{sheet_name}' has different data."
    
    return True, "All sheets and cells are identical."



